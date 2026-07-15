import telebot
import requests
import pandas as pd
import re
import threading
import time
from telebot import types
import io
import os  # OS মডিউল নিশ্চিত করার জন্য যুক্ত করা হলো
from openpyxl import Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

import os
BOT_TOKEN = os.getenv("BOT_TOKEN")
bot = telebot.TeleBot(BOT_TOKEN)

lock = threading.Lock()
processing = False

# কনকারেন্ট থ্রেড সংখ্যা (একসাথে কতগুলো অ্যাকাউন্ট চেক হবে)
# স্পিড আরও বাড়াতে চাইলে ১০ থেকে বাড়িয়ে ১৫ বা ২০ করতে পারেন।
MAX_WORKERS = 10 

# =========================================
# PARSE FOR CHECKING LOGIC ONLY
# =========================================
def parse_account_from_row(row_values):
    if len(row_values) < 4:
        return None
    
    name = str(row_values[0]).strip()
    email = str(row_values[1]).strip()
    password = str(row_values[2]).strip()
    
    raw_d = str(row_values[3]).strip()
    d_parts = raw_d.split("|")
    
    if len(d_parts) >= 4:
        client_id = d_parts[-1]
        refresh_token = "|".join(d_parts[2:-1])
        if not re.match(r'^[0-9a-fA-F-]{36}$', client_id):
            refresh_token = "|".join(d_parts[2:])
            client_id = ""
    else:
        return None
    
    return {
        "name": name,
        "email": email,
        "password": password,
        "refresh_token": refresh_token,
        "client_id": client_id
    }

def parse_single_input(input_text):
    parts = input_text.split("|")
    parts = [p.strip() for p in parts]
    
    if len(parts) >= 7:
        email = parts[3]
        password = parts[2]
        client_id = parts[-1]
        refresh_token = "|".join(parts[5:-1])
        return {"email": email, "password": password, "refresh_token": refresh_token, "client_id": client_id}
    elif len(parts) >= 4:
        return {"email": parts[0], "password": parts[1], "refresh_token": parts[2], "client_id": parts[3]}
    else:
        return None

# =========================================
# GET ACCESS TOKEN
# =========================================
def get_access_token(refresh_token, client_id):
    url = "https://login.microsoftonline.com/common/oauth2/v2.0/token"
    data = {
        "client_id": client_id,
        "grant_type": "refresh_token",
        "refresh_token": refresh_token,
        "scope": "offline_access Mail.Read"
    }
    try:
        r = requests.post(url, data=data, timeout=20)
        if r.status_code == 200:
            return r.json().get("access_token"), None
        else:
            error_json = r.json()
            error_desc = error_json.get("error_description", "No description")
            error_code = error_json.get("error", "Unknown")
            return None, f"{error_code}: {error_desc[:300]}"
    except Exception as e:
        return None, str(e)

# =========================================
# GET FIRST EMAIL
# =========================================
def get_first_mail_subject_and_body(access_token):
    url = "https://graph.microsoft.com/v1.0/me/messages?$top=1&$orderby=receivedDateTime desc&$select=subject,bodyPreview"
    headers = {"Authorization": f"Bearer {access_token}"}
    try:
        r = requests.get(url, headers=headers, timeout=15)
        if r.status_code == 200:
            mails = r.json().get("value", [])
            if mails:
                return mails[0].get("subject", ""), mails[0].get("bodyPreview", ""), None
            return "", "", None
        else:
            return None, None, f"Graph API Error {r.status_code}"
    except Exception as e:
        return None, None, str(e)

# =========================================
# CLASSIFY
# =========================================
def classify_account(subject, body_preview=""):
    if not subject and not body_preview:
        return "NO_AMA"
    
    sub_lower = (subject or "").lower()
    body_lower = (body_preview or "").lower()
    
    verify_patterns = [
        "verify your new amazon account",
        "verify your amazon account",
        "amazon account verification",
        "welcome to amazon",
        "verify your email address for amazon"
    ]
    
    suspended_patterns = [
        "take action",
        "suspended",
        "suspendido",
        "se ha suspendido",
        "your account has been suspended",
        "cuenta de amazon ha sido suspendida"
    ]
    
    combined = sub_lower + " " + body_lower
    
    for pattern in verify_patterns:
        if pattern in combined:
            return "LIVE"
    
    has_suspended = any(p in combined for p in suspended_patterns)
    if has_suspended:
        return "DEAD"
    
    return "NO_AMA"

# =========================================
# READ XLSX (KEEPS ORIGINAL ROWS)
# =========================================
def read_accounts_from_xlsx(file_bytes):
    accounts = []
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), engine='openpyxl', header=None)
        for idx, row in df.iterrows():
            row_values = [str(v) if pd.notna(v) else "" for v in row]
            if len(row_values) < 4:
                continue
            
            acc_parsed = parse_account_from_row(row_values)
            if acc_parsed:
                accounts.append({
                    "parsed": acc_parsed,
                    "original_row": row_values
                })
    except Exception as e:
        print(f"XL read error: {e}")
    return accounts

# =========================================
# EXPORT TO EXCEL (ONLY ORIGINAL COLUMNS)
# =========================================
def export_to_excel(live_accounts, dead_accounts, no_amazon_accounts, token_failed_accounts, api_failed_accounts):
    wb_good = Workbook()
    ws_good = wb_good.active
    ws_good.title = "GOOD_LIVE"
    
    for row_idx, original_row in enumerate(live_accounts, 1):
        for col_idx, val in enumerate(original_row, 1):
            ws_good.cell(row=row_idx, column=col_idx, value=val)
            
    good_bytes = io.BytesIO()
    wb_good.save(good_bytes)
    good_bytes.seek(0)
    
    wb_bad = Workbook()
    ws_bad = wb_bad.active
    ws_bad.title = "BAD"
    
    all_bad_rows = dead_accounts + no_amazon_accounts + token_failed_accounts + api_failed_accounts
    
    for row_idx, original_row in enumerate(all_bad_rows, 1):
        for col_idx, val in enumerate(original_row, 1):
            ws_bad.cell(row=row_idx, column=col_idx, value=val)
            
    bad_bytes = io.BytesIO()
    wb_bad.save(bad_bytes)
    bad_bytes.seek(0)
    
    return good_bytes, bad_bytes

# =========================================
# CHECK SINGLE ACCOUNT (WORKER TASK)
# =========================================
def worker_check_task(item):
    acc = item["parsed"]
    orig = item["original_row"]
    
    result, subject, body_preview, error = check_single_account(
        acc['email'], acc['password'], acc['refresh_token'], acc['client_id']
    )
    return result, orig, acc['email']

def check_single_account(email, password, refresh_token, client_id):
    access_token, token_error = get_access_token(refresh_token, client_id)
    if not access_token:
        return "TOKEN_FAIL", None, None, token_error
    
    subject, body_preview, api_error = get_first_mail_subject_and_body(access_token)
    if api_error:
        return "API_ERROR", None, None, api_error
    
    result = classify_account(subject, body_preview)
    return result, subject, body_preview, None

# =========================================
# REPLY KEYBOARD ONLY (WITH NEW LOGO)
# =========================================
def get_reply_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
    # বাটন এ শপিং ব্যাগ ও বক্স ইমোজি দিয়ে লোগো ভাইব দেওয়া হয়েছে
    markup.add(types.KeyboardButton("📦 Amazon Submit 🛍️"))
    return markup

# =========================================
# START COMMAND
# =========================================
@bot.message_handler(commands=["start"])
def start(message):
    markup = get_reply_keyboard()
    bot.send_photo(
        message.chat.id,
        "https://upload.wikimedia.org/wikipedia/commons/thumb/a/a9/Amazon_logo.svg/2560px-Amazon_logo.svg.png",
        caption=(
            "⚡ *Turbo Fast Amazon Checker by MaF!*\n\n"
            "👋 স্বাগতম! আপনার এক্সেল ফাইলটি সরাসরি চ্যাটে ড্রপ করুন অথবা নিচের বাটনে ক্লিক করুন।"
        ),
        parse_mode="Markdown",
        reply_markup=markup
    )

# =========================================
# MULTI-THREADED EXCEL PROCESSING
# =========================================
@bot.message_handler(content_types=["document"])
def handle_file(message):
    global processing
    
    if processing:
        bot.reply_to(message, "⏳ একটি ফাইল অলরেডি প্রোসেস হচ্ছে। অনুগ্রহ করে অপেক্ষা করুন...")
        return
    
    file_name = message.document.file_name
    if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
        bot.reply_to(message, "❌ শুধু বৈধ `.xlsx` বা `.xls` ফাইল পাঠান!")
        return
    
    processing = True
    status_msg = bot.reply_to(message, f"📥 রিড করা হচ্ছে `{file_name}`...")
    
    def master_worker():
        global processing
        try:
            file_info = bot.get_file(message.document.file_id)
            downloaded = bot.download_file(file_info.file_path)
            
            accounts = read_accounts_from_xlsx(downloaded)
            
            if not accounts:
                bot.edit_message_text(
                    "❌ ফাইলে কোনো সঠিক ডাটা বা অ্যাকাউন্ট পাওয়া যায়নি!",
                    chat_id=message.chat.id,
                    message_id=status_msg.message_id
                )
                processing = False
                return
            
            total_accs = len(accounts)
            bot.edit_message_text(
                f"🚀 মোট {total_accs}টি অ্যাকাউন্ট পাওয়া গেছে। ফাস্ট চেকিং শুরু হচ্ছে...",
                chat_id=message.chat.id,
                message_id=status_msg.message_id
            )
            
            live_accounts = []
            dead_accounts = []
            no_amazon_accounts = []
            token_failed_accounts = []
            api_failed_accounts = []
            
            completed_count = 0
            last_update_time = time.time()
            
            # ThreadPoolExecutor দিয়ে প্যারালাল (Fast) প্রসেসিং
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                future_to_acc = {executor.submit(worker_check_task, item): item for item in accounts}
                
                for future in as_completed(future_to_acc):
                    result, orig, email = future.result()
                    completed_count += 1
                    
                    if result == "TOKEN_FAIL":
                        token_failed_accounts.append(orig)
                    elif result == "API_ERROR":
                        api_failed_accounts.append(orig)
                    elif result == "LIVE":
                        live_accounts.append(orig)
                    elif result == "DEAD":
                        dead_accounts.append(orig)
                    else:
                        no_amazon_accounts.append(orig)
                    
                    # প্রতি ৩টি অ্যাকাউন্ট পর পর অথবা ১.৫ সেকেন্ড পর পর লাইভ আপডেট দেবে (টেলিগ্রাম লিমিট এড়াতে)
                    if completed_count % 3 == 0 or (time.time() - last_update_time) > 1.5:
                        try:
                            bot.edit_message_text(
                                f"⚡ Chaking : [ {completed_count} / {total_accs} ]\n"
                                f"✅ Live: {len(live_accounts)} | ❌ Dead: {len(dead_accounts)} | ⚠️ No Ama: {len(no_amazon_accounts)}",
                                chat_id=message.chat.id,
                                message_id=status_msg.message_id
                            )
                            last_update_time = time.time()
                        except:
                            pass
            
            summary = (
                f"📊 *FAST CHECK COMPLETE*\n"
                f"━━━━━━━━━━━━━━━\n"
                f"📁 Total Rows: {total_accs}\n"
                f"✅ LIVE: {len(live_accounts)}\n"
                f"❌ DEAD: {len(dead_accounts)}\n"
                f"⚠️ NO AMA: {len(no_amazon_accounts)}\n"
                f"💥 Token Fail: {len(token_failed_accounts)}\n"
                f"🌐 API Error: {len(api_failed_accounts)}"
            )
            
            good_bytes, bad_bytes = export_to_excel(
                live_accounts, dead_accounts, 
                no_amazon_accounts, token_failed_accounts, 
                api_failed_accounts
            )
            
            if live_accounts:
                good_bytes.name = "GOOD_LIVE.xlsx"
                bot.send_document(
                    message.chat.id, good_bytes,
                    caption=f"✅ GOOD - {len(live_accounts)} LIVE accounts"
                )
            
            total_bad = len(dead_accounts) + len(no_amazon_accounts) + len(token_failed_accounts) + len(api_failed_accounts)
            if total_bad > 0:
                bad_bytes.name = "BAD_DEAD_FAILED.xlsx"
                bot.send_document(
                    message.chat.id, bad_bytes,
                    caption=f"❌ BAD - {total_bad} accounts (DEAD + FAILED)"
                )
            
            bot.edit_message_text(
                summary + "\n\n📊 অরিজিনাল ডাটা সহ ফাইল পাঠানো কমপ্লিট!",
                chat_id=message.chat.id,
                message_id=status_msg.message_id,
                parse_mode="Markdown"
            )
        
        except Exception as e:
            try:
                bot.edit_message_text(
                    f"❌ Error: {str(e)[:500]}",
                    chat_id=message.chat.id,
                    message_id=status_msg.message_id
                )
            except:
                pass
        
        finally:
            processing = False
            
    threading.Thread(target=master_worker).start()

# =========================================
# TEXT MESSAGE & KEYBOARD CLICK HANDLER
# =========================================
@bot.message_handler(func=lambda m: True)
def handle_text(message):
    # নতুন লোগো দেওয়া টেক্সট ম্যাচিং
    if "Amazon Submit" in message.text:
        bot.reply_to(
            message,
            "📤 *Import Excel File Instruction*\n\n"
            "আপনার এক্সেল ফাইলটি সরাসরি এই চ্যাটে অ্যাটাচমেন্ট হিসেবে দিন।\n\n"
            "ফরম্যাট লেআউট:\n"
            "A=Name | B=Email | C=Pass | D=email\\|user\\|RT\\|ClientID\n\n"
            "ফলাফলে আপনার পাঠানো অরিজিনাল কলামগুলো হুবহু ফেরত দেওয়া হবে।",
            parse_mode="Markdown"
        )
    elif message.text.startswith("/check"):
        try:
            input_text = message.text.split("/check", 1)[1].strip()
            acc_data = parse_single_input(input_text)
            
            if not acc_data:
                bot.reply_to(message, "❌ Format: `/check email|pass|refresh_token|client_id`", parse_mode="Markdown")
                return
                
            email = acc_data["email"]
            password = acc_data["password"]
            refresh_token = acc_data["refresh_token"]
            client_id = acc_data["client_id"]
            
            msg = bot.reply_to(message, f"🔍 Checking `{email}`...", parse_mode="Markdown")
            
            def worker():
                result, subject, body_preview, error = check_single_account(email, password, refresh_token, client_id)
                emoji = {"LIVE": "✅", "DEAD": "❌", "NO_AMA": "⚠️", "TOKEN_FAIL": "💥", "API_ERROR": "🌐"}
                bot.edit_message_text(
                    f"📧 *Account Check Result:*\n\n"
                    f"🏷️ Status: {emoji.get(result, '❓')} *{result}*\n"
                    f"📬 Email: `{email}`",
                    chat_id=message.chat.id,
                    message_id=msg.message_id,
                    parse_mode="Markdown"
                )
            threading.Thread(target=worker).start()
        except Exception as e:
            bot.reply_to(message, f"❌ Error: {str(e)}")
    else:
        markup = get_reply_keyboard()
        bot.reply_to(
            message,
            "🤖 *Amazon Checker by MaF!*\n\n"
            "গাইডলাইন দেখতে নিচে থাকা `📦 Amazon Submit 🛍️` বাটনে প্রেস করুন অথবা সরাসরি ফাইল সেন্ড করুন।",
            reply_markup=markup
        )

# =========================================
# RUN BOT
# =========================================
# ---------- UI ----------
os.system("cls")
os.system("title MaFi Shadow ⚡ Secure Console")

print("🚀 Turbo Fast Amazon Checker by MaF! RUNNING")
bot.infinity_polling(skip_pending=True, timeout=30)
